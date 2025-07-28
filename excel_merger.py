import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import json
import os
from pathlib import Path
import threading
from datetime import datetime, timezone, timedelta
from models import Item, Format, Product, Shipping

class ExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("엑셀 파일 통합 프로그램")
        self.root.geometry("800x600")
        
        # 업로드된 파일들을 저장할 딕셔너리
        self.uploaded_files = {
            "쌀": {"네이버": None, "올웨이즈": None},
            "구운란": {"쿠팡": None, "올웨이즈": None, "토스": None}
        }
        
        self.setup_ui()
    
    def setup_ui(self):
        """UI를 설정합니다."""
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 제목
        title_label = ttk.Label(main_frame, text="엑셀 파일 통합 프로그램", 
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # 탭 컨트롤
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 20))
        
        # 쌀 탭
        self.rice_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.rice_frame, text="쌀")
        self.setup_rice_tab()
        
        # 구운란 탭
        self.egg_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.egg_frame, text="구운란")
        self.setup_egg_tab()
        
        # 통합 버튼
        merge_button = ttk.Button(main_frame, text="파일 통합하기", 
                                 command=self.merge_files, style="Accent.TButton")
        merge_button.grid(row=2, column=0, columnspan=2, pady=20)
        
        # 상태 표시
        self.status_label = ttk.Label(main_frame, text="파일을 업로드해주세요.", 
                                     font=("Arial", 10))
        self.status_label.grid(row=3, column=0, columnspan=2)
        
        # 그리드 가중치 설정
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
    def setup_rice_tab(self):
        """쌀 탭을 설정합니다."""
        # 네이버 섹션
        naver_frame = ttk.LabelFrame(self.rice_frame, text="네이버", padding="10")
        naver_frame.grid(row=0, column=0, padx=10, pady=10, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.naver_label = ttk.Label(naver_frame, text="파일을 선택하세요")
        self.naver_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        naver_button = ttk.Button(naver_frame, text="파일 선택", 
                                 command=lambda: self.select_file("쌀", "네이버"))
        naver_button.grid(row=1, column=0, padx=(0, 5))
        
        naver_clear = ttk.Button(naver_frame, text="초기화", 
                                command=lambda: self.clear_file("쌀", "네이버"))
        naver_clear.grid(row=1, column=1)
        
        # 올웨이즈 섹션
        always_frame = ttk.LabelFrame(self.rice_frame, text="올웨이즈", padding="10")
        always_frame.grid(row=0, column=1, padx=10, pady=10, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.always_label = ttk.Label(always_frame, text="파일을 선택하세요")
        self.always_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        always_button = ttk.Button(always_frame, text="파일 선택", 
                                  command=lambda: self.select_file("쌀", "올웨이즈"))
        always_button.grid(row=1, column=0, padx=(0, 5))
        
        always_clear = ttk.Button(always_frame, text="초기화", 
                                 command=lambda: self.clear_file("쌀", "올웨이즈"))
        always_clear.grid(row=1, column=1)
        
        # 그리드 가중치 설정
        self.rice_frame.columnconfigure(0, weight=1)
        self.rice_frame.columnconfigure(1, weight=1)
        self.rice_frame.rowconfigure(0, weight=1)
        
    def setup_egg_tab(self):
        """구운란 탭을 설정합니다."""
        # 쿠팡 섹션
        coupang_frame = ttk.LabelFrame(self.egg_frame, text="쿠팡", padding="10")
        coupang_frame.grid(row=0, column=0, padx=10, pady=10, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.coupang_label = ttk.Label(coupang_frame, text="파일을 선택하세요")
        self.coupang_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        coupang_button = ttk.Button(coupang_frame, text="파일 선택", 
                                   command=lambda: self.select_file("구운란", "쿠팡"))
        coupang_button.grid(row=1, column=0, padx=(0, 5))
        
        coupang_clear = ttk.Button(coupang_frame, text="초기화", 
                                  command=lambda: self.clear_file("구운란", "쿠팡"))
        coupang_clear.grid(row=1, column=1)
        
        # 올웨이즈 섹션
        always_egg_frame = ttk.LabelFrame(self.egg_frame, text="올웨이즈", padding="10")
        always_egg_frame.grid(row=0, column=1, padx=10, pady=10, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.always_egg_label = ttk.Label(always_egg_frame, text="파일을 선택하세요")
        self.always_egg_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        always_egg_button = ttk.Button(always_egg_frame, text="파일 선택", 
                                      command=lambda: self.select_file("구운란", "올웨이즈"))
        always_egg_button.grid(row=1, column=0, padx=(0, 5))
        
        always_egg_clear = ttk.Button(always_egg_frame, text="초기화", 
                                     command=lambda: self.clear_file("구운란", "올웨이즈"))
        always_egg_clear.grid(row=1, column=1)
        
        # 토스 섹션
        toss_frame = ttk.LabelFrame(self.egg_frame, text="토스", padding="10")
        toss_frame.grid(row=0, column=2, padx=10, pady=10, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.toss_label = ttk.Label(toss_frame, text="파일을 선택하세요")
        self.toss_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        toss_button = ttk.Button(toss_frame, text="파일 선택", 
                                command=lambda: self.select_file("구운란", "토스"))
        toss_button.grid(row=1, column=0, padx=(0, 5))
        
        toss_clear = ttk.Button(toss_frame, text="초기화", 
                               command=lambda: self.clear_file("구운란", "토스"))
        toss_clear.grid(row=1, column=1)
        
        # 그리드 가중치 설정
        self.egg_frame.columnconfigure(0, weight=1)
        self.egg_frame.columnconfigure(1, weight=1)
        self.egg_frame.columnconfigure(2, weight=1)
        self.egg_frame.rowconfigure(0, weight=1)
    

    
    def select_file(self, category, company):
        """파일 선택 다이얼로그를 엽니다."""
        file_path = filedialog.askopenfilename(
            title=f"{company} 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.upload_file(category, company, file_path)
    
    def upload_file(self, category, company, file_path):
        """파일을 업로드합니다."""
        try:
            # 파일 존재 확인
            if not os.path.exists(file_path):
                messagebox.showerror("오류", "파일을 찾을 수 없습니다.")
                return
            
            # 엑셀 파일 읽기 테스트
            df = pd.read_excel(file_path)
            if df.empty:
                messagebox.showerror("오류", "빈 엑셀 파일입니다.")
                return
            
            # 파일 저장
            self.uploaded_files[category][company] = file_path
            
            # UI 업데이트
            self.update_file_label(category, company, os.path.basename(file_path))
            
            # 상태 업데이트
            self.update_status()
            
        except Exception as e:
            messagebox.showerror("오류", f"파일을 읽는 중 오류가 발생했습니다: {str(e)}")
    
    def clear_file(self, category, company):
        """업로드된 파일을 초기화합니다."""
        self.uploaded_files[category][company] = None
        self.update_file_label(category, company, "파일을 선택하세요")
        self.update_status()
    
    def update_file_label(self, category, company, filename):
        """파일 라벨을 업데이트합니다."""
        label_map = {
            ("쌀", "네이버"): self.naver_label,
            ("쌀", "올웨이즈"): self.always_label,
            ("구운란", "쿠팡"): self.coupang_label,
            ("구운란", "올웨이즈"): self.always_egg_label,
            ("구운란", "토스"): self.toss_label
        }
        
        label = label_map.get((category, company))
        if label:
            label.config(text=filename)
    
    def update_status(self):
        """상태를 업데이트합니다."""
        current_tab = self.notebook.select()
        if current_tab == str(self.rice_frame):
            uploaded_count = sum(1 for f in self.uploaded_files["쌀"].values() if f is not None)
            total_count = len(self.uploaded_files["쌀"])
            self.status_label.config(text=f"쌀: {uploaded_count}/{total_count} 파일 업로드됨")
        elif current_tab == str(self.egg_frame):
            uploaded_count = sum(1 for f in self.uploaded_files["구운란"].values() if f is not None)
            total_count = len(self.uploaded_files["구운란"])
            self.status_label.config(text=f"구운란: {uploaded_count}/{total_count} 파일 업로드됨")
    
    def merge_files(self):
        """파일들을 통합합니다."""
        current_tab = self.notebook.select()
        
        if current_tab == str(self.rice_frame):
            category = "쌀"
        elif current_tab == str(self.egg_frame):
            category = "구운란"
        else:
            return
        
        # 업로드된 파일 확인
        uploaded_files = [f for f in self.uploaded_files[category].values() if f is not None]
        if not uploaded_files:
            messagebox.showwarning("경고", "업로드된 파일이 없습니다.")
            return
        
        # 통합 작업을 별도 스레드에서 실행
        threading.Thread(target=self._merge_files_thread, args=(category,), daemon=True).start()
    
    def _merge_files_thread(self, category):
        """별도 스레드에서 파일 통합을 수행합니다."""
        try:
            if category == "구운란":
                self._process_egg_files()
            else:
                self._process_rice_files()
                    
        except Exception as e:
            # 오류 메시지 표시 (메인 스레드에서)
            print(e)
            self.root.after(0, lambda: messagebox.showerror("오류", f"파일 통합 중 오류가 발생했습니다: {str(e)}"))
    
    def _process_egg_files(self):
        """구운란 파일들을 처리합니다."""
        items = []
        
        platform_files = {}
        if self.uploaded_files['구운란'].get('쿠팡', None) is not None:
            platform_files['coupang'] = self.uploaded_files['구운란']['쿠팡']
        if self.uploaded_files['구운란'].get('올웨이즈', None) is not None:
            platform_files['alwayz'] = self.uploaded_files['구운란']['올웨이즈']
        if self.uploaded_files['구운란'].get('토스', None) is not None:
            platform_files['toss'] = self.uploaded_files['구운란']['토스']
        
        for platform, file_path in platform_files.items():
            print(platform, file_path)
            if file_path is not None:
                try:
                    df = pd.read_excel(file_path)
                    extracted_items = self.extract_data(df, platform, 'egg')
                    items.extend(extracted_items)
                except Exception as e:
                    print(e)
                    # self.root.after(0, lambda: messagebox.showerror("오류", f"{platform} 파일 처리 중 오류: {str(e)}"))
                    return
        
        if items:
            self.export_excel(items)
        else:
            self.root.after(0, lambda: messagebox.showwarning("경고", "처리할 데이터가 없습니다."))
    
    def extract_data(self, df, platform, product_type):
        """데이터를 추출하여 Item 객체들을 반환합니다."""
        # Product['egg'][platform]이 리스트인지 확인하고 필터링

        filtered_df = df[df[Format['product_id'][platform]].isin(Product[product_type][platform])]
        
        # 그룹화된 데이터로 Item 객체들 생성
        items = []
        if product_type == 'egg':
            for _, row in filtered_df.iterrows():
                receiver = row[Format['receiver'][platform]]
                contact = row[Format['contact'][platform]]
                address = row[Format['address'][platform]]
                message = row[Format['message'][platform]]
                count = row[Format['count'][platform]]
                
                if platform == 'coupang':
                    product_name = row[Format['product_name'][platform]].split()[1]
                    quantity = int(row[Format['options'][platform]].split()[1][:2]) * row[Format['count'][platform]]
                elif platform == 'alwayz':
                    split_options = row[Format['options'][platform]].split('\n')
                    product_name = split_options[0].split()[2]
                    quantity = int(split_options[1].split()[2][:2]) * row[Format['count'][platform]]
                    contact_str = str(row[Format['contact'][platform]])
                    contact = f'0{contact_str[:-8]}-{contact_str[-8:-4]}-{contact_str[-4:]}'
                elif platform == 'toss':
                    product_name = row[Format['product_name'][platform]].split()[-1].split('(')[0]
                    quantity = int(row[Format['options'][platform]].split(',')[0][:2]) * row[Format['count'][platform]]
                    contact_str = str(row[Format['contact'][platform]])
                    contact = f'0{contact_str[:-8]}-{contact_str[-8:-4]}-{contact_str[-4:]}'
                
                if quantity > 120:
                    _quantity = quantity
                    while _quantity > 0:
                        if _quantity > 120:
                            item = Item(
                                receiver=receiver,
                                contact=contact,
                                address=address,
                                quantity=f'{product_name} 120구',
                                note=message,
                                shipping_number=Shipping[platform],
                            )
                            items.append(item)
                            _quantity -= 120
                        else:
                            item = Item(
                                receiver=receiver,
                                contact=contact,
                                address=address,
                                quantity=f'{product_name} {_quantity}구',
                                note=message,
                                shipping_number=Shipping[platform],
                            )
                            items.append(item)
                            break
                else:
                    item = Item(
                        receiver=receiver,
                        contact=contact,
                        address=address,
                        quantity=f'{product_name} {quantity}구',
                        note=message,
                        shipping_number=Shipping[platform],
                    )
                    items.append(item)
            
            return items
        elif product_type == 'rice':
            for _, row in filtered_df.iterrows():
                receiver = row[Format['receiver'][platform]]
                contact = row[Format['contact'][platform]]
                address = row[Format['address'][platform]]
                message = row[Format['message'][platform]]
                count = row[Format['count'][platform]]
                unit = 'kg'
                
                if platform == 'naver':
                    product_name = row[Format['product_name'][platform]].split()[1]
                    quantity = int(row[Format['options'][platform]].split()[1][:2]) * row[Format['count'][platform]]
                elif platform == 'alwayz':
                    if row[Format['product_code'][platform]] == '백미':
                        product_name = '백미팝(150g)'
                        quantity = 5 * row[Format['count'][platform]]
                        unit = '개'
                    elif row[Format['product_code'][platform]] == '신4':
                        product_name = '신동진'
                        quantity = 4 * row[Format['count'][platform]]
                        unit = 'kg'
                    elif row[Format['product_code'][platform]] == '신10':
                        product_name = '신동진'
                        quantity = 10 * row[Format['count'][platform]]
                        unit = 'kg'
                    contact_str = str(row[Format['contact'][platform]])
                    contact = f'0{contact_str[:-8]}-{contact_str[-8:-4]}-{contact_str[-4:]}'

                if unit == 'kg':
                    if quantity > 20:
                        _quantity = quantity
                        while _quantity > 0:
                            if _quantity > 20:
                                item = Item(
                                    receiver=receiver,
                                    contact=contact,
                                    address=address,
                                    quantity=f'20{unit}',
                                    note=message,
                                    shipping_number=Shipping[platform],
                                )
                                items.append(item)
                                _quantity -= 20
                            else:
                                item = Item(
                                    receiver=receiver,
                                    contact=contact,
                                    address=address,
                                    quantity=f'{_quantity}{unit}',
                                    note=message,
                                    shipping_number=Shipping[platform],
                                )
                                items.append(item)
                                break
                    else:
                        item = Item(
                        receiver=receiver,
                        contact=contact,
                        address=address,
                        quantity=f'{quantity}{unit}',
                        note=message,
                        shipping_number=Shipping[platform],
                    )
                    items.append(item)
                    
                else:
                    item = Item(
                        receiver=receiver,
                        contact=contact,
                        address=address,
                        quantity=f'{product_name} {quantity}{unit}',
                        note=message,
                        shipping_number=Shipping[platform],
                    )
                    items.append(item)
            
            return items
            

        
    
    def export_excel(self, items):
        """Item 객체들을 엑셀 파일로 저장합니다."""
        # items 리스트를 엑셀 파일로 저장
        excel_data = []
        for item in items:
            row = [
                item.receiver,           # 받으시는분
                item.contact,            # 연락처
                item.receiver_manager,   # 받으시는분담당자
                item.receiver_phone,     # 받으시는분핸드폰
                item.zip_code,           # 우편번호
                item.address,            # 총주소
                item.quantity,           # 수량
                item.item_name,          # 품목명
                item.freight,            # 운임
                item.payment_condition,  # 지불조건
                item.shipping_number,    # 출고번호
                item.note,               # 특기사항
                "",                      # 빈 칼럼
                item.tracking_number,    # 운송장번호
            ]
            excel_data.append(row)
        
        # DataFrame 생성 및 엑셀 저장
        columns = [
            '받으시는분', '연락처', '받으시는분담당자', '받으시는분핸드폰', '우편번호', 
            '총주소', '수량', '품목명', '운임', '지불조건', '출고번호', '특기사항', '', '운송장번호'
        ]
        
        result_df = pd.DataFrame(excel_data, columns=columns)
        
        # KST로 오늘 날짜를 YYYYMMDD 형태로 저장
        kst = timezone(timedelta(hours=9))
        today = datetime.now(kst).strftime('%Y%m%d')
        
        # 맥북 데스크톱 경로 설정
        desktop_path = os.path.expanduser("~/Desktop")
        output_filename = f'(취합포맷)1.포커스양식_구운란_{today}.xlsx'
        output_path = os.path.join(desktop_path, output_filename)
        
        # ExcelWriter를 사용하여 스타일링 적용
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # 워크시트 가져오기
            worksheet = writer.sheets['Sheet1']
            
            # 스타일링을 위한 import
            from openpyxl.styles import PatternFill, Alignment
            
            # 정렬 설정
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # 노란색 배경 설정
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # 헤더 행(1행)에 노란색 배경과 왼쪽 정렬 적용
            for cell in worksheet[1]:
                cell.fill = yellow_fill
                cell.alignment = left_alignment
            
            # 데이터 행들(2행부터)에 왼쪽 정렬 적용
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = left_alignment
            
            # 자동 필터 적용 (헤더 행 포함)
            worksheet.auto_filter.ref = f"A1:{chr(ord('A') + len(columns) - 1)}{worksheet.max_row}"
            
            # 칼럼 너비 설정
            column_widths = {
                'A': 15,  # 받으시는분
                'B': 15,  # 연락처
                'C': 17,  # 받으시는분담당자
                'D': 17,  # 받으시는분핸드폰
                'E': 10,  # 우편번호
                'F': 60,  # 총주소
                'G': 10,  # 수량
                'H': 10,  # 품목명
                'I': 10,  # 운임
                'J': 10,  # 지불조건
                'K': 10,  # 출고번호
                'L': 25,  # 특기사항
                'M': 5,   # 빈 칼럼
                'N': 10,  # 운송장번호
            }
            
            for column, width in column_widths.items():
                worksheet.column_dimensions[column].width = width
        
        # UI 업데이트 (메인 스레드에서)
        self.root.after(0, lambda: messagebox.showinfo("완료", f"파일이 성공적으로 생성되었습니다.\n파일명: {output_filename}"))
    
    def _process_rice_files(self):
        """쌀 파일들을 처리합니다."""
        items = []

        platform_files = {}
        if self.uploaded_files['쌀'].get('네이버', None) is not None:
            platform_files['coupang'] = self.uploaded_files['쌀']['네이버']
        if self.uploaded_files['쌀'].get('올웨이즈', None) is not None:
            platform_files['alwayz'] = self.uploaded_files['쌀']['올웨이즈']
        
        for platform, file_path in platform_files.items():
            if file_path is not None:
                try:
                    df = pd.read_excel(file_path)
                    extracted_items = self.extract_data(df, platform, 'rice')
                    items.extend(extracted_items)
                except Exception as e:
                    print(e)
                    # self.root.after(0, lambda: messagebox.showerror("오류", f"{platform} 파일 처리 중 오류: {str(e)}"))
                    return
        
        if items:
            self.export_excel(items)
        else:
            self.root.after(0, lambda: messagebox.showwarning("경고", "처리할 데이터가 없습니다."))
    
    
def main():
    root = tk.Tk()
    app = ExcelMergerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main() 