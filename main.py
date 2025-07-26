#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import excel_merger
import pandas as pd
from models import Item, Format, Product, Shipping
# def main():
#     excel_merger.main()

def extract_data(df, platform):
    # Product['egg'][platform]이 리스트인지 확인하고 필터링
    filtered_df = df[df[Format['product_id'][platform]].isin(Product['egg'][platform])]
    
    # 수취인전화번호와 등록상품명으로 그룹화하여 수량 합산
    # grouped_df = filtered_df.groupby(['수취인전화번호', '등록상품명']).agg({
    #     '구매수(수량)': 'sum',
    #     '수취인이름': 'first',
    #     '등록옵션명': 'first',
    #     '수취인 주소': 'first',
    #     '구매수(수량)': 'first',
    #     '배송메세지': 'first',
    #     '등록옵션명': 'first',
    # }).reset_index()

    # 그룹화된 데이터로 Item 객체들 생성
    items = []
    for _, row in filtered_df.iterrows():
        receiver = row[Format['receiver'][platform]]
        contact = row[Format['contact'][platform]]
        address = row[Format['address'][platform]]
        message = row[Format['message'][platform]]
        count = row[Format['count'][platform]]
        
        if platform == 'coupang':
            product_name = row[Format['product_name'][platform]].split()[1]
            quantity = int(row[Format['options'][platform]].split()[1][:2]) * row[Format['count'][platform]]
        elif platform == 'alwayz':
            split_options = row[Format['options'][platform]].split('\n')
            product_name = split_options[0].split()[2]
            quantity = int(split_options[1].split()[2][:2]) * row[Format['count'][platform]]
            contact_str = str(row[Format['contact'][platform]])
            contact =f'0{contact_str[:-8]}-{contact_str[-8:-4]}-{contact_str[-4:]}'
        elif platform == 'toss':
            product_name = row[Format['product_name'][platform]].split()[-1].split('(')[0]
            quantity = int(row[Format['options'][platform]].split(',')[0][:2]) * row[Format['count'][platform]]
            contact_str = str(row[Format['contact'][platform]])
            contact =f'0{contact_str[:-8]}-{contact_str[-8:-4]}-{contact_str[-4:]}'

        
        if quantity > 120:
            _quantity = quantity
            while True:
                if _quantity < 240:
                    for _ in range(2):
                        item = Item(
                            receiver=receiver,
                            contact=contact,
                            address=address,
                            quantity=f'{product_name} {_quantity//2}구',
                            note=message,
                            shipping_number=Shipping[platform],
                        )
                        items.append(item)
                    break
                else:
                    item = Item(
                        receiver=receiver,
                        contact=contact,
                        address=address,
                        quantity=f'{product_name} 120구',
                        note=message,
                        shipping_number=Shipping[platform],
                    )
                    items.append(item)
                    _quantity -= 120
        else:
            item = Item(
                receiver=receiver,
                contact=contact,
                address=address,
                quantity=f'{product_name} {quantity}구',
                note=message,
                shipping_number=Shipping[platform],
            )
            items.append(item)
    return items

def export_excel(items):
    # items 리스트를 엑셀 파일로 저장
    excel_data = []
    for item in items:
        row = [
            item.receiver,           # 받으시는분
            item.contact,            # 연락처
            item.receiver_manager,   # 받으시는분담당자
            item.receiver_phone,     # 받으시는분핸드폰
            item.zip_code,           # 우편번호
            item.address,            # 총주소
            item.quantity,           # 수량
            item.item_name,          # 품목명
            item.freight,            # 운임
            item.payment_condition,  # 지불조건
            item.shipping_number,    # 출고번호
            item.note,               # 특기사항
            "",                      # 빈 칼럼
            item.tracking_number,    # 운송장번호
        ]
        excel_data.append(row)
    
    # DataFrame 생성 및 엑셀 저장
    columns = [
        '받으시는분', '연락처', '받으시는분담당자', '받으시는분핸드폰', '우편번호', 
        '총주소', '수량', '품목명', '운임', '지불조건', '출고번호', '특기사항', '', '운송장번호'
    ]
    
    result_df = pd.DataFrame(excel_data, columns=columns)
    
    # KST로 오늘 날짜를 YYYYMMDD 형태로 저장
    from datetime import datetime, timezone, timedelta
    kst = timezone(timedelta(hours=9))
    today = datetime.now(kst).strftime('%Y%m%d')
    
    output_filename = f'(취합포맷)1.포커스양식_구운란_{today}.xlsx'
    
    # ExcelWriter를 사용하여 스타일링 적용
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # 워크시트 가져오기
        worksheet = writer.sheets['Sheet1']
        
        # 스타일링을 위한 import
        from openpyxl.styles import PatternFill, Alignment
        
        # 정렬 설정
        left_alignment = Alignment(horizontal='left', vertical='center')
        # 노란색 배경 설정
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # 헤더 행(1행)에 노란색 배경과 왼쪽 정렬 적용
        for cell in worksheet[1]:
            cell.fill = yellow_fill
            cell.alignment = left_alignment
        
        # 데이터 행들(2행부터)에 오른쪽 정렬 적용
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = left_alignment
        
        # 자동 필터 적용 (헤더 행 포함)
        worksheet.auto_filter.ref = f"A1:{chr(ord('A') + len(columns) - 1)}{worksheet.max_row}"
        
        # 칼럼 너비 설정
        column_widths = {
            'A': 15,  # 받으시는분
            'B': 15,  # 연락처
            'C': 17,  # 받으시는분담당자
            'D': 17,  # 받으시는분핸드폰
            'E': 10,  # 우편번호
            'F': 60,  # 총주소
            'G': 10,  # 수량
            'H': 10,  # 품목명
            'I': 10,  # 운임
            'J': 10,  # 지불조건
            'K': 10,  # 출고번호
            'L': 25,  # 특기사항
            'M': 5,   # 빈 칼럼
            'N': 10,  # 운송장번호
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
    
    print(f"\n엑셀 파일이 '{output_filename}'로 저장되었습니다.")

def main():
    items = []
    df_coupang = pd.read_excel('쿠팡.xlsx')
    # platform = 'coupang'

    df_alwayz = pd.read_excel('올웨이즈.xlsx')
    # platform = 'alwayz'

    df_toss = pd.read_excel('토스.xlsx')
    # platform = 'toss'

    items += extract_data(df_coupang, 'coupang')
    items += extract_data(df_alwayz, 'alwayz')
    items += extract_data(df_toss, 'toss')
    export_excel(items)    
    
    print("헤더가 노란색으로 설정되고, 필터와 칼럼 너비가 적용되었습니다.")

if __name__ == "__main__":
    main()
